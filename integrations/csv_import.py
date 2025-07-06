"""
CSV/Excel Import Service for Kitchentory
Handles bulk import of inventory items from CSV and Excel files.
"""

import csv
import io
import logging
import re
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass

import openpyxl
from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from django.utils import timezone

from inventory.models import Category, Product, StorageLocation, InventoryItem
from inventory.services import product_data_service
from .models import ImportJob, ImportStatus, ImportSource, ParsedReceiptItem

logger = logging.getLogger(__name__)


@dataclass
class ImportMapping:
    """Defines how CSV columns map to inventory fields"""
    name: str = None
    brand: str = None
    quantity: str = None
    unit: str = None
    price: str = None
    category: str = None
    location: str = None
    expiration_date: str = None
    notes: str = None
    barcode: str = None


@dataclass
class ValidationError:
    """Represents a validation error for an import row"""
    row: int
    field: str
    value: str
    message: str


@dataclass
class ImportPreview:
    """Preview of import data with validation results"""
    total_rows: int
    valid_rows: int
    invalid_rows: int
    sample_data: List[Dict]
    errors: List[ValidationError]
    suggested_mappings: ImportMapping
    column_names: List[str]


class CSVImportService:
    """Service for importing inventory from CSV/Excel files"""
    
    # Supported file types
    SUPPORTED_EXTENSIONS = ['.csv', '.xlsx', '.xls']
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
    MAX_ROWS = 5000  # Reasonable limit for processing
    
    # Common column name mappings
    COLUMN_MAPPINGS = {
        'name': ['name', 'product', 'item', 'product_name', 'item_name', 'title'],
        'brand': ['brand', 'manufacturer', 'make', 'brand_name'],
        'quantity': ['quantity', 'qty', 'amount', 'count', 'number'],
        'unit': ['unit', 'units', 'measurement', 'uom', 'unit_of_measure'],
        'price': ['price', 'cost', 'amount', 'value', 'price_paid'],
        'category': ['category', 'type', 'group', 'section', 'department'],
        'location': ['location', 'storage', 'place', 'where', 'storage_location'],
        'expiration_date': ['expiration', 'expires', 'exp_date', 'expiry', 'best_by', 'use_by'],
        'notes': ['notes', 'note', 'description', 'comment', 'memo'],
        'barcode': ['barcode', 'upc', 'ean', 'code', 'product_code']
    }
    
    # Standard units for normalization
    UNIT_MAPPINGS = {
        'item': ['item', 'items', 'each', 'ea', 'piece', 'pieces', 'pcs'],
        'oz': ['oz', 'ounce', 'ounces'],
        'lb': ['lb', 'lbs', 'pound', 'pounds'],
        'g': ['g', 'gram', 'grams'],
        'kg': ['kg', 'kilogram', 'kilograms'],
        'ml': ['ml', 'milliliter', 'milliliters'],
        'l': ['l', 'liter', 'liters', 'litre', 'litres'],
        'fl_oz': ['fl oz', 'fl. oz', 'fluid ounce', 'fluid ounces'],
        'cup': ['cup', 'cups', 'c'],
        'tbsp': ['tbsp', 'tablespoon', 'tablespoons'],
        'tsp': ['tsp', 'teaspoon', 'teaspoons'],
        'pack': ['pack', 'package', 'pkg'],
        'count': ['count', 'ct'],
    }

    def __init__(self, household, user):
        self.household = household
        self.user = user

    def validate_file(self, file: UploadedFile) -> Tuple[bool, str]:
        """Validate uploaded file before processing"""
        
        # Check file extension
        file_ext = self._get_file_extension(file.name)
        if file_ext not in self.SUPPORTED_EXTENSIONS:
            return False, f"Unsupported file type. Supported formats: {', '.join(self.SUPPORTED_EXTENSIONS)}"
        
        # Check file size
        if file.size > self.MAX_FILE_SIZE:
            return False, f"File too large. Maximum size: {self.MAX_FILE_SIZE // (1024*1024)}MB"
        
        # Check if file is readable
        try:
            file.seek(0)
            if file_ext == '.csv':
                # Try to read first few lines as CSV
                content = file.read(1024).decode('utf-8')
                file.seek(0)
                csv.Sniffer().sniff(content)
            else:
                # Try to open as Excel
                workbook = openpyxl.load_workbook(file, read_only=True)
                workbook.close()
                file.seek(0)
                
        except Exception as e:
            return False, f"Invalid file format: {str(e)}"
        
        return True, "File is valid"

    def preview_import(self, file: UploadedFile, mapping: ImportMapping = None) -> ImportPreview:
        """Preview import data and suggest column mappings"""
        
        try:
            # Parse file data
            rows, headers = self._parse_file(file)
            
            # Auto-detect column mappings if not provided
            if mapping is None:
                mapping = self._auto_detect_mappings(headers)
            
            # Validate sample of rows
            sample_size = min(100, len(rows))
            sample_rows = rows[:sample_size]
            
            valid_count = 0
            errors = []
            sample_data = []
            
            for i, row in enumerate(sample_rows):
                try:
                    validated_item = self._validate_row(row, mapping, headers, i + 1)
                    if validated_item:
                        valid_count += 1
                        sample_data.append(validated_item)
                except Exception as e:
                    errors.append(ValidationError(
                        row=i + 1,
                        field='general',
                        value='',
                        message=str(e)
                    ))
            
            return ImportPreview(
                total_rows=len(rows),
                valid_rows=valid_count,
                invalid_rows=len(sample_rows) - valid_count,
                sample_data=sample_data[:10],  # Only show first 10 for preview
                errors=errors[:20],  # Limit errors shown
                suggested_mappings=mapping,
                column_names=headers
            )
            
        except Exception as e:
            logger.error(f"Error previewing import: {e}")
            raise ValueError(f"Could not preview file: {str(e)}")

    def process_import(self, file: UploadedFile, mapping: ImportMapping) -> ImportJob:
        """Process the full import with the given mapping"""
        
        # Create import job
        import_job = ImportJob.objects.create(
            household=self.household,
            user=self.user,
            source=ImportSource.CSV_UPLOAD,
            status=ImportStatus.PROCESSING,
            file_path=file.name
        )
        
        try:
            # Parse file data
            rows, headers = self._parse_file(file)
            import_job.total_items = len(rows)
            import_job.save()
            
            # Process all rows
            valid_items = []
            errors = []
            
            for i, row in enumerate(rows):
                try:
                    validated_item = self._validate_row(row, mapping, headers, i + 1)
                    if validated_item:
                        valid_items.append(validated_item)
                    else:
                        errors.append({
                            'row': i + 1,
                            'error': 'Validation failed'
                        })
                except Exception as e:
                    errors.append({
                        'row': i + 1,
                        'error': str(e)
                    })
            
            # Store validated data
            import_job.validated_data = valid_items
            import_job.errors = errors
            import_job.processed_items = len(valid_items)
            import_job.failed_items = len(errors)
            import_job.save()
            
            # Create inventory items
            created_count = self._create_inventory_items(import_job, valid_items)
            
            # Update job status
            import_job.created_items = created_count
            import_job.status = ImportStatus.COMPLETED
            import_job.completed_at = timezone.now()
            import_job.save()
            
            logger.info(f"Import job {import_job.id} completed: {created_count} items created")
            
            return import_job
            
        except Exception as e:
            logger.error(f"Error processing import job {import_job.id}: {e}")
            import_job.status = ImportStatus.FAILED
            import_job.errors = [{'general': str(e)}]
            import_job.save()
            raise

    def _parse_file(self, file: UploadedFile) -> Tuple[List[List[str]], List[str]]:
        """Parse CSV or Excel file and return rows and headers"""
        
        file_ext = self._get_file_extension(file.name)
        file.seek(0)
        
        if file_ext == '.csv':
            return self._parse_csv(file)
        else:
            return self._parse_excel(file)

    def _parse_csv(self, file: UploadedFile) -> Tuple[List[List[str]], List[str]]:
        """Parse CSV file"""
        
        # Try different encodings
        encodings = ['utf-8', 'utf-8-sig', 'latin1', 'cp1252']
        
        for encoding in encodings:
            try:
                file.seek(0)
                content = file.read().decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError("Could not decode file with any supported encoding")
        
        # Detect CSV dialect
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(content[:1024])
        
        # Parse CSV
        reader = csv.reader(io.StringIO(content), dialect=dialect)
        rows = list(reader)
        
        if not rows:
            raise ValueError("CSV file is empty")
        
        headers = [self._normalize_header(h) for h in rows[0]]
        data_rows = rows[1:]
        
        if len(data_rows) > self.MAX_ROWS:
            raise ValueError(f"Too many rows. Maximum allowed: {self.MAX_ROWS}")
        
        return data_rows, headers

    def _parse_excel(self, file: UploadedFile) -> Tuple[List[List[str]], List[str]]:
        """Parse Excel file"""
        
        try:
            workbook = openpyxl.load_workbook(file, read_only=True, data_only=True)
            worksheet = workbook.active
            
            rows = []
            for row in worksheet.iter_rows(values_only=True):
                # Convert None values to empty strings and all values to strings
                row_data = [str(cell) if cell is not None else '' for cell in row]
                rows.append(row_data)
            
            workbook.close()
            
            if not rows:
                raise ValueError("Excel file is empty")
            
            headers = [self._normalize_header(h) for h in rows[0]]
            data_rows = rows[1:]
            
            if len(data_rows) > self.MAX_ROWS:
                raise ValueError(f"Too many rows. Maximum allowed: {self.MAX_ROWS}")
            
            return data_rows, headers
            
        except Exception as e:
            raise ValueError(f"Could not read Excel file: {str(e)}")

    def _normalize_header(self, header: str) -> str:
        """Normalize header name for mapping"""
        return re.sub(r'[^a-zA-Z0-9_]', '_', str(header).lower().strip())

    def _auto_detect_mappings(self, headers: List[str]) -> ImportMapping:
        """Auto-detect column mappings based on header names"""
        
        mapping = ImportMapping()
        
        for header in headers:
            normalized = header.lower()
            
            # Check each field mapping
            for field, keywords in self.COLUMN_MAPPINGS.items():
                if any(keyword in normalized for keyword in keywords):
                    setattr(mapping, field, header)
                    break
        
        return mapping

    def _validate_row(self, row: List[str], mapping: ImportMapping, 
                     headers: List[str], row_num: int) -> Optional[Dict]:
        """Validate a single row and return normalized data"""
        
        # Create mapping from headers to values
        if len(row) != len(headers):
            raise ValueError(f"Row {row_num}: Column count mismatch")
        
        row_data = dict(zip(headers, row))
        
        # Extract and validate required fields
        name = self._get_mapped_value(row_data, mapping.name)
        if not name or not name.strip():
            return None  # Skip rows without names
        
        # Validate and normalize other fields
        validated = {
            'name': name.strip(),
            'brand': self._get_mapped_value(row_data, mapping.brand, '').strip(),
            'quantity': self._validate_quantity(
                self._get_mapped_value(row_data, mapping.quantity, '1')
            ),
            'unit': self._normalize_unit(
                self._get_mapped_value(row_data, mapping.unit, 'item')
            ),
            'price': self._validate_price(
                self._get_mapped_value(row_data, mapping.price, '')
            ),
            'category': self._get_mapped_value(row_data, mapping.category, '').strip(),
            'location': self._get_mapped_value(row_data, mapping.location, '').strip(),
            'expiration_date': self._validate_date(
                self._get_mapped_value(row_data, mapping.expiration_date, '')
            ),
            'notes': self._get_mapped_value(row_data, mapping.notes, '').strip(),
            'barcode': self._get_mapped_value(row_data, mapping.barcode, '').strip(),
            'row_number': row_num
        }
        
        return validated

    def _get_mapped_value(self, row_data: Dict, mapping: str, default: str = '') -> str:
        """Get value from row data using mapping"""
        if not mapping or mapping not in row_data:
            return default
        return str(row_data[mapping]).strip()

    def _validate_quantity(self, value: str) -> Decimal:
        """Validate and convert quantity to Decimal"""
        if not value:
            return Decimal('1')
        
        try:
            # Clean up the value
            cleaned = re.sub(r'[^\d.]', '', str(value))
            if not cleaned:
                return Decimal('1')
            
            quantity = Decimal(cleaned)
            if quantity <= 0:
                return Decimal('1')
            return quantity
            
        except (InvalidOperation, ValueError):
            return Decimal('1')

    def _validate_price(self, value: str) -> Optional[Decimal]:
        """Validate and convert price to Decimal"""
        if not value:
            return None
        
        try:
            # Clean up the value (remove currency symbols, etc.)
            cleaned = re.sub(r'[^\d.]', '', str(value))
            if not cleaned:
                return None
            
            price = Decimal(cleaned)
            return price if price >= 0 else None
            
        except (InvalidOperation, ValueError):
            return None

    def _validate_date(self, value: str) -> Optional[date]:
        """Validate and convert date string to date object"""
        if not value or str(value).lower() in ['', 'none', 'null', 'n/a']:
            return None
        
        # Common date formats to try
        date_formats = [
            '%Y-%m-%d',
            '%m/%d/%Y',
            '%d/%m/%Y',
            '%m-%d-%Y',
            '%d-%m-%Y',
            '%Y/%m/%d',
            '%B %d, %Y',
            '%b %d, %Y',
            '%d %B %Y',
            '%d %b %Y'
        ]
        
        value_str = str(value).strip()
        
        for fmt in date_formats:
            try:
                return datetime.strptime(value_str, fmt).date()
            except ValueError:
                continue
        
        return None

    def _normalize_unit(self, value: str) -> str:
        """Normalize unit to standard format"""
        if not value:
            return 'item'
        
        normalized = value.lower().strip()
        
        for standard_unit, variations in self.UNIT_MAPPINGS.items():
            if normalized in variations:
                return standard_unit
        
        return normalized if normalized else 'item'

    def _create_inventory_items(self, import_job: ImportJob, valid_items: List[Dict]) -> int:
        """Create inventory items from validated data"""
        
        created_count = 0
        
        with transaction.atomic():
            for item_data in valid_items:
                try:
                    # Get or create storage location
                    location = None
                    if item_data.get('location'):
                        location, _ = StorageLocation.objects.get_or_create(
                            household=self.household,
                            name=item_data['location'],
                            defaults={'location_type': 'other'}
                        )
                    
                    # Try to find existing product by name and brand
                    product = None
                    if item_data.get('barcode'):
                        try:
                            product = Product.objects.get(barcode=item_data['barcode'])
                        except Product.DoesNotExist:
                            pass
                    
                    if not product:
                        # Create or get product
                        product, created = Product.objects.get_or_create(
                            name=item_data['name'],
                            brand=item_data.get('brand', ''),
                            defaults={
                                'barcode': item_data.get('barcode', ''),
                                'default_unit': item_data['unit'],
                                'category': self._get_or_create_category(item_data.get('category'))
                            }
                        )
                    
                    # Create inventory item
                    inventory_item = InventoryItem.objects.create(
                        household=self.household,
                        product=product,
                        quantity=item_data['quantity'],
                        unit=item_data['unit'],
                        price_paid=item_data.get('price'),
                        expiration_date=item_data.get('expiration_date'),
                        location=location,
                        notes=f"Imported from CSV. {item_data.get('notes', '')}".strip(),
                        added_by=self.user
                    )
                    
                    # Create parsed receipt item for tracking
                    ParsedReceiptItem.objects.create(
                        import_job=import_job,
                        name=item_data['name'],
                        brand=item_data.get('brand', ''),
                        quantity=item_data['quantity'],
                        unit=item_data['unit'],
                        price=item_data.get('price'),
                        is_processed=True,
                        is_approved=True,
                        created_inventory_item_id=inventory_item.id,
                        line_number=item_data.get('row_number')
                    )
                    
                    created_count += 1
                    
                except Exception as e:
                    logger.error(f"Error creating inventory item from import: {e}")
                    continue
        
        return created_count

    def _get_or_create_category(self, category_name: str) -> Optional['Category']:
        """Get or create category by name"""
        if not category_name:
            return None
        
        try:
            return Category.objects.get(name__iexact=category_name)
        except Category.DoesNotExist:
            # Create new category
            return Category.objects.create(
                name=category_name.title(),
                slug=category_name.lower().replace(' ', '-')
            )

    def _get_file_extension(self, filename: str) -> str:
        """Get file extension from filename"""
        return '.' + filename.split('.')[-1].lower() if '.' in filename else ''

    @classmethod
    def get_sample_csv(cls) -> str:
        """Generate a sample CSV file content for download"""
        
        sample_data = [
            ['name', 'brand', 'quantity', 'unit', 'price', 'category', 'location', 'expiration_date', 'notes'],
            ['Organic Bananas', 'Fresh Market', '2', 'lb', '3.99', 'Produce', 'Fridge', '2024-01-15', 'Ripe and sweet'],
            ['Whole Milk', 'Dairy Farm', '1', 'gal', '4.49', 'Dairy', 'Fridge', '2024-01-10', ''],
            ['Chicken Breast', 'Premium Poultry', '2.5', 'lb', '12.99', 'Meat', 'Freezer', '2024-02-01', 'Boneless, skinless'],
            ['Brown Rice', 'Healthy Grains', '5', 'lb', '6.99', 'Pantry', 'Pantry', '', 'Long grain'],
            ['Olive Oil', 'Mediterranean', '1', 'bottle', '8.99', 'Pantry', 'Pantry', '2025-06-01', 'Extra virgin']
        ]
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerows(sample_data)
        
        return output.getvalue()